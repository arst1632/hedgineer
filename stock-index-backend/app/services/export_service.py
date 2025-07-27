import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any
import io
from app.services.index_service import index_service

class ExportService:
    def __init__(self):
        pass
    
    def export_to_excel(self, start_date: str, end_date: str) -> bytes:
        """Export all index data to Excel"""
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Get data
        performance_data = index_service.get_index_performance(start_date, end_date)
        composition_changes = index_service.get_composition_changes(start_date, end_date)
        
        # Create Performance sheet
        self._create_performance_sheet(wb, performance_data)
        
        # Create Composition Changes sheet
        self._create_changes_sheet(wb, composition_changes)
        
        # Create daily compositions sheet (sample dates)
        self._create_compositions_sheet(wb, start_date, end_date)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    def _create_performance_sheet(self, wb: Workbook, data: list):
        """Create performance sheet"""
        ws = wb.create_sheet("Index Performance")
        
        # Headers
        headers = ["Date", "Daily Return (%)", "Cumulative Return (%)", "Index Value"]
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for record in data:
            ws.append([
                record['date'],
                round(record['daily_return'] * 100, 4),
                round(record['cumulative_return'] * 100, 4),
                round(record['index_value'], 2)
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_changes_sheet(self, wb: Workbook, data: list):
        """Create composition changes sheet"""
        ws = wb.create_sheet("Composition Changes")
        
        # Headers
        headers = ["Date", "Stocks Entered", "Stocks Exited", "Total Changes"]
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for record in data:
            ws.append([
                record['date'],
                ", ".join(record['entered']) if record['entered'] else "None",
                ", ".join(record['exited']) if record['exited'] else "None",
                record['total_changes']
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_compositions_sheet(self, wb: Workbook, start_date: str, end_date: str):
        """Create sample compositions sheet"""
        ws = wb.create_sheet("Sample Compositions")
        
        # Get a few sample dates
        performance_data = index_service.get_index_performance(start_date, end_date)
        if not performance_data:
            return
        
        # Take first, middle, and last dates
        sample_dates = [performance_data[0]['date']]
        if len(performance_data) > 2:
            mid_idx = len(performance_data) // 2
            sample_dates.append(performance_data[mid_idx]['date'])
        if len(performance_data) > 1:
            sample_dates.append(performance_data[-1]['date'])
        
        # Headers
        headers = ["Date", "Rank", "Symbol", "Weight (%)", "Market Cap", "Price"]
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add composition data for sample dates
        for date in sample_dates:
            composition = index_service.get_index_composition(date)
            for stock in composition[:10]:  # Show top 10 for each date
                ws.append([
                    date,
                    stock['rank'],
                    stock['symbol'],
                    round(stock['weight'] * 100, 2),
                    f"${stock['market_cap']:,.0f}",
                    f"${stock['price']:.2f}"
                ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

export_service = ExportService()